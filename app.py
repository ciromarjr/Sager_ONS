#Codigo baixar planilhas
# -*- encoding: utf-8 -*-
from datetime import datetime
import os
import requests
import time
from requests.exceptions import RequestException
import openpyxl

# Configurações da API
auth_url = "https://pops.ons.org.br/ons.pop.federation/oauth2/token"
exportar_url = "https://integra.ons.org.br/api/sager/renovaveis/api/RelatorioApuracaoUsinasRenovaveis/exportarRelatorio"
api_auth_payload = {
    "grant_type": "password",
    "client_id": "SAGER",
    "username": "",  # Substitua pelo seu usuário
    "password": ""  # Substitua pela sua senha
}

# Dicionários fornecidos
tipoRelatorio = {
    1: "Relatório Geral",
    2: "Relatório Geração de Referência"
}

# Mapeamento dos IDs para os números de série que serão usados no nome dos arquivos
numero_serie = {
    00: 00,   # Codigo do Sager e do numero do arquivo da planilha
    
}

idsConjuntos = {
    00: "Conjunto", # Codigo do Sager e do conjunto
    
}

usuario = os.getenv('USERNAME')  # Ou use os.getlogin()
# Caminho base onde estão salvas as planilhas
caminho2 = fr"C:\Users\{usuario}\Relatórios\SAGER"
caminho1 = fr"C:\Users\{usuario}\Relatórios\SAGER"

# Função para verificar qual caminho existe
def verificar_caminho_existente(caminho1, caminho2):
    if os.path.exists(caminho1):
        return caminho1
    elif os.path.exists(caminho2):
        return caminho2
    else:
        return None

# Verifica qual caminho existe
caminho = verificar_caminho_existente(caminho1, caminho2)

if caminho is None:
    print("Nenhum dos caminhos existe. Verifique os caminhos fornecidos.")
else:
    print(f"O caminho existente é: {caminho}")


# Função para obter o token de autenticação da API
def get_api_token():
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded',
        'Origin': 'https://apps18.ons.org.br'
    }
    response = requests.post(auth_url, data=api_auth_payload, headers=headers)
    if response.status_code == 200:
        auth_data = response.json()
        return auth_data['access_token']
    else:
        print("Erro na autenticação da API:", response.status_code, response.text)
        return None

# Função para determinar e criar as pastas base
def criar_pastas_base():
    ano_atual = datetime.now().year
    base_dir = os.path.join(caminho, f"{ano_atual}_SAGER")

    # Verifica se as pastas principais existem, se não, as cria
    relatorio_geral_dir = os.path.join(base_dir, "Relatório Geral")
    relatorio_geracao_referencia_dir = os.path.join(base_dir, "Relatório Geração de Referência")

    if not os.path.exists(relatorio_geral_dir):
        os.makedirs(relatorio_geral_dir)
    if not os.path.exists(relatorio_geracao_referencia_dir):
        os.makedirs(relatorio_geracao_referencia_dir)
    
    return relatorio_geral_dir, relatorio_geracao_referencia_dir

# Função para baixar o relatório e salvá-lo na pasta correta
def baixar_relatorio(token, tipo_relatorio_id, conjunto_id, conjunto_nome, output_dir, ano):
    headers = {
        'Authorization': f'Bearer {token}',
        'Accept': 'application/json',
        'Content-Type': 'application/json',
        'User-Agent': 'Mozilla/5.0',
        'Origin': 'https://apps18.ons.org.br'
    }

    # Obter a data de início e fim para o ano atual
    data_inicio = f"{ano}-01-01T00:00:00.000Z"
    data_fim = f"{ano}-12-31T23:59:59.000Z"

    # Corpo da solicitação
    body = {
        "dataInicio": data_inicio,
        "dataFim": data_fim,
        "apenasAgentesFavoritos": False,
        "tipoRelatorio": tipo_relatorio_id,
        "periodoBase": None,
        "idsFontes": [],
        "idsAgentes": [],
        "idsUsinas": [],
        "idsConjuntos": [conjunto_id]
    }

    try:
        # Faz a requisição ao servidor
        response = requests.post(exportar_url, headers=headers, json=body)

        if response.status_code == 200:
            # Verifica se há conteúdo na resposta
            if len(response.content) > 0:
                # Nome do relatório com base no tipo, nome do conjunto e ano
                relatorio_tipo = "RelatorioGeral" if tipo_relatorio_id == 1 else "RelatorioGeracaoReferencia"
                numero = numero_serie[conjunto_id]
                filename = f"{relatorio_tipo}_{numero}_{conjunto_nome.replace(' ', '_')}_{ano}.xlsx"
                
                # Caminho completo para salvar o arquivo
                file_path = os.path.join(output_dir, filename)
                
                # Salva o relatório no arquivo
                with open(file_path, "wb") as file:
                    file.write(response.content)
                print(f"Relatório de {ano} baixado com sucesso! Salvo em: {file_path}")
            else:
                print(f"Sem dados disponíveis para {conjunto_nome} em {ano}.")
        else:
            print(f"Erro ao baixar o relatório de {ano} para {conjunto_nome}: {response.status_code} - {response.text}.")

    except RequestException as e:
        # Em caso de erro, exibe a mensagem e passa para o próximo conjunto/ano
        print(f"Erro de requisição ao tentar baixar o relatório de {ano} para {conjunto_nome}: {str(e)}.")

# Função para processar uma usina por vez
def processar_usina(conjunto_id):
    # Obter o token de autenticação da API
    token = get_api_token()
    
    if token:
        conjunto_nome = idsConjuntos[conjunto_id]
        relatorio_geral_dir, relatorio_geracao_referencia_dir = criar_pastas_base()
        
        for tipo_relatorio_id, tipo_relatorio_nome in tipoRelatorio.items():
            # Determinar a pasta de saída
            output_dir = relatorio_geral_dir if tipo_relatorio_id == 1 else relatorio_geracao_referencia_dir

            # Iterar sobre os anos e baixar os relatórios anuais
            ano_inicial = 2024  # Defina o primeiro ano disponível
            ano_atual = datetime.now().year
            for ano in range(ano_inicial, ano_atual + 1):
                # Baixar o relatório e salvá-lo na pasta determinada
                baixar_relatorio(token, tipo_relatorio_id, conjunto_id, conjunto_nome, output_dir, ano)

# Função para mesclar os dados das planilhas anuais em uma única planilha
def mesclar_planilhas(conjunto_id):
    conjunto_nome = idsConjuntos[conjunto_id]
    numero = numero_serie[conjunto_id]
    
    # Diretório onde estão os relatórios gerais e os de geração de referência das usinas
    caminho_planilhas_gerais = os.path.join(caminho, f"{datetime.now().year}_SAGER", "Relatório Geral")
    caminho_planilhas_referencia = os.path.join(caminho, f"{datetime.now().year}_SAGER", "Relatório Geração de Referência")
    
    caminho_planilha_geral = os.path.join(caminho_planilhas_gerais, f"RelatorioGeral_{numero}_{conjunto_nome.replace(' ', '_')}.xlsx")
    
    workbook_geral = openpyxl.Workbook()
    
    # Criação das abas na planilha geral
    aba_patamares_geral = workbook_geral.create_sheet("Patamares")
    aba_restricoes_geral = workbook_geral.create_sheet("Restrições")
    aba_referencia_geral = workbook_geral.create_sheet("Geração Referência")
    
    # Variáveis de controle para manter o cabeçalho da primeira planilha
    cabecalho_patamares = True
    cabecalho_restricoes = True
    cabecalho_referencia = True

    # Mescla os dados de todas as planilhas anuais disponíveis
    for ano in range(2025, datetime.now().year + 1):   #Escolha 
        # Mesclando as planilhas gerais
        filename_geral = f"RelatorioGeral_{numero}_{conjunto_nome.replace(' ', '_')}_{ano}.xlsx"
        file_path_geral = os.path.join(caminho_planilhas_gerais, filename_geral)
        
        if os.path.exists(file_path_geral):
            workbook = openpyxl.load_workbook(file_path_geral, data_only=True)
            
            # Copiando dados da aba 'Patamares' a partir da linha 11
            if 'Patamares' in workbook.sheetnames:
                aba_patamares = workbook['Patamares']
                if cabecalho_patamares:
                    # Copiar o cabeçalho da primeira planilha
                    for row in aba_patamares.iter_rows(min_row=1, max_row=10, values_only=True):
                        aba_patamares_geral.append(row)
                    cabecalho_patamares = False  # Cabeçalho já copiado, não copiar novamente
                # Copiar dados abaixo do cabeçalho
                for row in aba_patamares.iter_rows(min_row=11, values_only=True):
                    aba_patamares_geral.append(row)
            
            # Copiando dados da aba 'Restrições'
            if 'Restrições' in workbook.sheetnames:
                aba_restricoes = workbook['Restrições']
                if cabecalho_restricoes:
                    # Copiar o cabeçalho completo da primeira planilha
                    for row in aba_restricoes.iter_rows(min_row=1, max_row=9, values_only=True):
                        aba_restricoes_geral.append(row)
                    cabecalho_restricoes = False  # Cabeçalho já copiado
                # Copiar dados abaixo do cabeçalho a partir da linha 10
                for row in aba_restricoes.iter_rows(min_row=10, values_only=True):
                    if any(cell is not None for cell in row):  # Verificar se a linha não está vazia
                        aba_restricoes_geral.append(row)
        
        # Mesclando as planilhas de referência
        filename_referencia = f"RelatorioGeracaoReferencia_{numero}_{conjunto_nome.replace(' ', '_')}_{ano}.xlsx"
        file_path_referencia = os.path.join(caminho_planilhas_referencia, filename_referencia)
        
        if os.path.exists(file_path_referencia):
            workbook_referencia = openpyxl.load_workbook(file_path_referencia, data_only=True)
            
            # Copiando dados da aba 'Geração Referência' a partir da linha 10
            if 'Geração Referência' in workbook_referencia.sheetnames:
                aba_referencia = workbook_referencia['Geração Referência']
                if cabecalho_referencia:
                    # Copiar o cabeçalho da primeira planilha
                    for row in aba_referencia.iter_rows(min_row=1, max_row=9, values_only=True):
                        aba_referencia_geral.append(row)
                    cabecalho_referencia = False  # Cabeçalho já copiado
                # Copiar dados abaixo do cabeçalho
                for row in aba_referencia.iter_rows(min_row=10, values_only=True):
                    if any(cell is not None for cell in row):  # Verificar se a linha não está vazia
                        aba_referencia_geral.append(row)

    # Remover a aba padrão criada pelo openpyxl se não tiver sido usada
    if 'Sheet' in workbook_geral.sheetnames:
        del workbook_geral['Sheet']
    
    # Salva a nova planilha geral com os dados mesclados
    workbook_geral.save(caminho_planilha_geral)
    print(f"Planilha geral criada com sucesso: {caminho_planilha_geral}")

# Execução principal
if __name__ == "__main__":
    for conjunto_id in idsConjuntos.keys():
        # Baixar os relatórios
        processar_usina(conjunto_id)
        
        # Mesclar os relatórios baixados
        mesclar_planilhas(conjunto_id)
